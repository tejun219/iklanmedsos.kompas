import http.server
import socketserver
import json
import os
import openpyxl
import re
import generate_data
import subprocess
import threading
import datetime
import glob

PORT = 8000
ADMIN_PASSWORD = "suroto2A"  # Secure admin password. Change as needed.
GOOGLE_SCRIPT_URL = ""  # Paste your deployed Google Apps Script Web App URL here to connect live to Google Sheets

MEMO_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "memo")

def auto_git_push():
    """Runs git commit and push in a background thread to automatically update the visitor dashboard on GitHub Pages."""
    # Find git executable
    git_exe = "git"
    username = os.environ.get("USERNAME", "")
    paths_to_check = [
        r"C:\Program Files\Git\bin\git.exe",
        r"C:\Program Files\Git\cmd\git.exe",
    ]
    if username:
        paths_to_check.append(rf"C:\Users\{username}\AppData\Local\GitHubDesktop\app-3.5.11\resources\app\git\cmd\git.exe")
        # Wildcard to find dynamic GitHub Desktop versions
        try:
            pattern = rf"C:\Users\{username}\AppData\Local\GitHubDesktop\app-*\resources\app\git\cmd\git.exe"
            matched = glob.glob(pattern)
            if matched:
                paths_to_check.extend(matched)
        except Exception as glob_err:
            print(f"[auto-git] Wildcard glob error: {glob_err}")

    for p in paths_to_check:
        if os.path.exists(p):
            git_exe = p
            break

    print(f"[auto-git] Menggunakan Git executable: {git_exe}")
    
    try:
        working_dir = os.path.dirname(os.path.abspath(__file__))
        
        # Git add
        add_cmd = [git_exe, "add", "data.js", "Iklan Media Sosial Harian Kompas.xlsx", "index.html"]
        subprocess.run(add_cmd, cwd=working_dir, check=True, capture_output=True, text=True)
        
        # Git commit
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        commit_msg = f"Auto-update outline data: {timestamp}"
        commit_cmd = [git_exe, "commit", "-m", commit_msg]
        subprocess.run(commit_cmd, cwd=working_dir, capture_output=True, text=True)
        
        # Determine branch to push
        branch_proc = subprocess.run([git_exe, "branch", "--show-current"], cwd=working_dir, capture_output=True, text=True)
        current_branch = branch_proc.stdout.strip() if branch_proc.returncode == 0 else "master"
        if not current_branch:
            current_branch = "master"
            
        push_ref = f"{current_branch}:main" if current_branch != "main" else "main"
        
        print(f"[auto-git] Memulai push {push_ref} ke origin...")
        push_cmd = [git_exe, "push", "origin", push_ref]
        push_res = subprocess.run(push_cmd, cwd=working_dir, check=True, capture_output=True, text=True)
        
        print(f"[auto-git] SUCCESS: Berhasil push update terbaru ke GitHub Pages!")
    except subprocess.CalledProcessError as e:
        print(f"[auto-git] ERROR: Perintah Git gagal dengan exit code {e.returncode}")
        print(f"Stdout: {e.stdout}")
        print(f"Stderr: {e.stderr}")
    except Exception as e:
        print(f"[auto-git] EXCEPTION: Gagal melakukan sinkronisasi otomatis ke GitHub: {e}")

def trigger_auto_git_push():
    """Triggers the background thread for Git push to keep UI responsive."""
    t = threading.Thread(target=auto_git_push)
    t.daemon = True
    t.start()


# Google Sheets Live Content Mirroring
# Spreadsheet publik berisi data Live Performance (Impressions, Reach, Engagement, Status, Link)
LIVE_CONTENT_SHEET_ID = "1rNc6Jb5lDE7PFdokZsHNUwqR8jitvb7dIiEWBPlW3lg"

# Mapping nama sheet (Bahasa Indonesia) ke GID di Google Sheets
# Tambahkan GID sesuai sheet yang ada di Google Sheets
LIVE_SHEET_GID_MAP = {
    "Januari 2026": None,
    "Februari 2026": None,
    "Maret 2026": None,
    "April 2026": None,
    "Mei 2026": None,
    "Juni 2026": "412501877",   # sheet aktif yang diketahui
    "Juli 2026": None,
}

def parse_google_sheets_url(url):
    if not url:
        return None, None, None
    id_match = re.search(r'/d/([a-zA-Z0-9_-]+)', url)
    gid_match = re.search(r'[#?&]gid=(\d+)', url)
    sheet_match = re.search(r'[?&]sheet=([^&]+)', url)
    
    spreadsheet_id = id_match.group(1) if id_match else None
    gid = gid_match.group(1) if gid_match else None
    sheet_name = urllib.parse.unquote(sheet_match.group(1)) if sheet_match else None
    
    return spreadsheet_id, gid, sheet_name

def fetch_live_content_csv(sheet_name):
    """Fetch CSV dari Google Sheets untuk sheet bulan tertentu. Return list of dicts."""
    import urllib.request
    import urllib.parse
    import csv
    import io
    
    # Ambil bagian bulan saja (misal "Juni 2026" -> "Juni")
    sheet_name_clean = sheet_name.split(" ")[0]
    
    # Check if there are dynamic KCM links configured
    target_sheet_id = LIVE_CONTENT_SHEET_ID
    target_gid = LIVE_SHEET_GID_MAP.get(sheet_name)
    target_sheet_name = sheet_name_clean
    
    if os.path.exists('kcm_links.json'):
        try:
            with open('kcm_links.json', 'r', encoding='utf-8') as f:
                links_map = json.load(f)
            month_link = links_map.get(sheet_name)
            if month_link:
                s_id, gid, s_name = parse_google_sheets_url(month_link)
                if s_id:
                    target_sheet_id = s_id
                if gid:
                    target_gid = gid
                if s_name:
                    target_sheet_name = s_name
        except Exception as e:
            print(f"[live-content] Gagal membaca kcm_links.json: {e}")
            
    # Try fetching by sheet name clean via Google Visualization API
    url_gviz = f"https://docs.google.com/spreadsheets/d/{target_sheet_id}/gviz/tq?tqx=out:csv&sheet={urllib.parse.quote(target_sheet_name)}"
    
    raw = None
    try:
        print(f"[live-content] Mencoba fetch sheet '{target_sheet_name}' dari ID '{target_sheet_id}' via gviz API: {url_gviz}")
        req = urllib.request.Request(url_gviz, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=10) as response:
            raw = response.read().decode("utf-8")
    except Exception as e:
        print(f"[live-content] Gagal fetch via gviz API: {e}")
        # Try fallback via GID if target_gid is available
        if target_gid:
            url_gid = f"https://docs.google.com/spreadsheets/d/{target_sheet_id}/export?format=csv&gid={target_gid}"
            print(f"[live-content] Mencoba fallback ke GID '{target_gid}': {url_gid}")
            try:
                req = urllib.request.Request(url_gid, headers={"User-Agent": "Mozilla/5.0"})
                with urllib.request.urlopen(req, timeout=10) as response:
                    raw = response.read().decode("utf-8-sig")
            except Exception as e2:
                print(f"[live-content] Gagal fetch via GID fallback: {e2}")
                return []
        else:
            return []

    if not raw:
        return []
        
    try:
        # Detect headers dynamically
        reader = csv.reader(io.StringIO(raw))
        rows_raw = list(reader)
        if not rows_raw:
            return []
            
        headers = [h.strip().lower() for h in rows_raw[0]]
        
        def get_index(names):
            for name in names:
                if name.lower() in headers:
                    return headers.index(name.lower())
            return -1
            
        no_idx = get_index(["no"])
        client_idx = get_index(["client"])
        tgl_idx = get_index(["tanggal post", "tanggal"])
        format_idx = get_index(["format"])
        platform_idx = get_index(["platform"])
        akun_idx = get_index(["akun"])
        status_idx = get_index(["status"])
        imp_idx = get_index(["impressions/views", "impressions", "impr"])
        reach_idx = get_index(["reach"])
        eng_idx = get_index(["engagement", "engage"])
        ctr_idx = get_index(["ctr (per impressions)", "ctr"])
        link_idx = get_index(["link konten", "link"])

        rows = []
        for r_cells in rows_raw[1:]:
            if not r_cells:
                continue
            
            client = r_cells[client_idx].strip() if client_idx != -1 and client_idx < len(r_cells) else ""
            tanggal = r_cells[tgl_idx].strip() if tgl_idx != -1 and tgl_idx < len(r_cells) else ""
            if not client and not tanggal:
                continue
            
            # Parse tanggal "4-Jun-26" → "2026-06-04"
            tgl_iso = ""
            if tanggal:
                try:
                    p = tanggal.split("-")
                    if len(p) == 3:
                        d_str = p[0].zfill(2)
                        m_str = p[1].lower()[:3]
                        y_str = p[2]
                        if len(y_str) == 2:
                            y_str = "20" + y_str
                        
                        month_map = {
                            "jan": "01", "feb": "02", "mar": "03", "apr": "04",
                            "may": "05", "mei": "05", "jun": "06", "jul": "07",
                            "aug": "08", "agu": "08", "sep": "09", "oct": "10",
                            "okt": "10", "nov": "11", "dec": "12", "des": "12"
                        }
                        mm = month_map.get(m_str, "01")
                        tgl_iso = f"{y_str}-{mm}-{d_str}"
                except:
                    tgl_iso = tanggal
            
            rows.append({
                "no":          r_cells[no_idx].strip() if no_idx != -1 and no_idx < len(r_cells) else "",
                "client":      client,
                "tgl_post":    tgl_iso,
                "format":      r_cells[format_idx].strip() if format_idx != -1 and format_idx < len(r_cells) else "",
                "platform":    r_cells[platform_idx].strip() if platform_idx != -1 and platform_idx < len(r_cells) else "",
                "akun":        r_cells[akun_idx].strip() if akun_idx != -1 and akun_idx < len(r_cells) else "",
                "status":      r_cells[status_idx].strip() if status_idx != -1 and status_idx < len(r_cells) else "",
                "impressions": r_cells[imp_idx].strip() if imp_idx != -1 and imp_idx < len(r_cells) else "",
                "reach":       r_cells[reach_idx].strip() if reach_idx != -1 and reach_idx < len(r_cells) else "",
                "engagement":  r_cells[eng_idx].strip() if eng_idx != -1 and eng_idx < len(r_cells) else "",
                "ctr":         r_cells[ctr_idx].strip() if ctr_idx != -1 and ctr_idx < len(r_cells) else "",
                "link":        r_cells[link_idx].strip() if link_idx != -1 and link_idx < len(r_cells) else "",
            })
        return rows
    except Exception as e:
        print(f"[live-content] Gagal memproses data CSV: {e}")
        return []


def fetch_all_live_sheet_gids():
    """Coba fetch daftar sheet dari Google Sheets (untuk auto-discover GID)."""
    import urllib.request
    import re as _re
    url = f"https://docs.google.com/spreadsheets/d/{LIVE_CONTENT_SHEET_ID}/edit"
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            html = resp.read().decode("utf-8", errors="ignore")
        # Cari pola name + gid dari HTML
        found = _re.findall(r'"name":"([^"]+)","index":\d+,"sheetId":(\d+)', html)
        return {name: str(gid) for name, gid in found}
    except Exception as e:
        print(f"[live-content] Gagal auto-discover sheet GIDs: {e}")
        return {}

def extract_memo_prefix(keterangan_order):
    if not keterangan_order:
        return ""
    # Find the first sequence of digits
    m = re.search(r'\b(\d+)\b', keterangan_order)
    if m:
        return m.group(1)
    # Fallback to any digits
    m = re.search(r'(\d+)', keterangan_order)
    if m:
        return m.group(1)
    return ""

def match_memo_file(prefix):
    if not prefix or not os.path.exists(MEMO_DIR):
        return None
    
    try:
        prefix_num = int(prefix)
    except ValueError:
        prefix_num = None
        
    files = os.listdir(MEMO_DIR)
    matching_files = []
    
    for f in files:
        file_prefix = extract_memo_prefix(f)
        if file_prefix:
            file_num = int(file_prefix) if file_prefix.isdigit() else None
            if prefix_num is not None and file_num == prefix_num:
                matching_files.append(f)
            elif file_prefix == prefix:
                matching_files.append(f)
                
    if not matching_files:
        return None
        
    # Sort by file modification time descending to get the latest memo file
    matching_files.sort(key=lambda x: os.path.getmtime(os.path.join(MEMO_DIR, x)), reverse=True)
    return os.path.join(MEMO_DIR, matching_files[0])

class CustomHandler(http.server.SimpleHTTPRequestHandler):
    def end_headers(self):
        # Prevent browser caching of static files and API responses
        self.send_header('Cache-Control', 'no-store, no-cache, must-revalidate, max-age=0')
        self.send_header('Pragma', 'no-cache')
        self.send_header('Expires', '0')
        super().end_headers()

    def handle_error(self, request, client_address):
        """Suppress benign connection errors (browser closing connection mid-transfer)."""
        pass

    def log_error(self, format, *args):
        """Suppress error log for connection aborted errors (WinError 10053/10054) — these are normal."""
        msg = format % args
        if '10053' in msg or '10054' in msg or 'ConnectionAbortedError' in msg or 'ConnectionResetError' in msg:
            return  # Ignore — browser closed the connection, not a real error
        super().log_error(format, *args)

    def log_visitor(self):
        import datetime
        ip = self.client_address[0]
        user_agent = self.headers.get('User-Agent', 'Unknown')
        visit_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        visitor_data = {
            "ip": ip,
            "time": visit_time,
            "user_agent": user_agent
        }
        
        try:
            visitors = []
            if os.path.exists('visitors.json'):
                with open('visitors.json', 'r', encoding='utf-8') as f:
                    try:
                        visitors = json.load(f)
                    except:
                        visitors = []
            
            # append and keep last 1000 records to avoid huge files
            visitors.append(visitor_data)
            if len(visitors) > 1000:
                visitors = visitors[-1000:]
                
            with open('visitors.json', 'w', encoding='utf-8') as f:
                json.dump(visitors, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Failed to log visitor: {e}")

    def do_POST(self):
        path_clean = self.path.split('?')[0]
        # 1. API: Login Admin password check
        if path_clean == '/api/login':
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            try:
                data = json.loads(post_data.decode('utf-8'))
                password = data.get('password', '')
                if password == ADMIN_PASSWORD:
                    self.send_response(200)
                    self.send_header('Content-Type', 'application/json')
                    self.end_headers()
                    self.wfile.write(json.dumps({"status": "success", "token": "admin-unlocked-token"}).encode('utf-8'))
                else:
                    self.send_response(401)
                    self.send_header('Content-Type', 'application/json')
                    self.end_headers()
                    self.wfile.write(json.dumps({"status": "error", "message": "wrong password"}).encode('utf-8'))
            except Exception as e:
                self.send_response(400)
                self.end_headers()
                self.wfile.write(str(e).encode('utf-8'))
                
        # 1.5 API: Sync data from Google Drive
        elif path_clean == '/api/sync':
            auth_header = self.headers.get('X-Admin-Password')
            if auth_header != ADMIN_PASSWORD:
                self.send_response(401)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({"status": "error", "message": "Akses Ditolak! Anda bukan Admin."}).encode('utf-8'))
                return
                
            try:
                # Force sync and rebuild data.js
                generate_data.main(sync=True)
                trigger_auto_git_push()
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({"status": "success", "message": "Berhasil sinkronisasi dengan Google Drive"}).encode('utf-8'))
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({"status": "error", "message": f"Server Error: {str(e)}"}).encode('utf-8'))
                
        # 1.6 API: Save Google Sheets KCM links dynamically
        elif path_clean == '/api/save-kcm-links':
            auth_header = self.headers.get('X-Admin-Password')
            if auth_header != ADMIN_PASSWORD:
                self.send_response(401)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({"status": "error", "message": "Akses Ditolak! Anda bukan Admin."}).encode('utf-8'))
                return
                
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            try:
                data = json.loads(post_data.decode('utf-8'))
                with open('kcm_links.json', 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({"status": "success", "message": "Berhasil menyimpan link Google Sheets Medsos Mirroring"}).encode('utf-8'))
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({"status": "error", "message": f"Server Error: {str(e)}"}).encode('utf-8'))

                
        # 2. API: Add new Ad Outline to Master Excel
        elif path_clean == '/api/add-ad':
            # Check Admin Authentication header
            auth_header = self.headers.get('X-Admin-Password')
            if auth_header != ADMIN_PASSWORD:
                self.send_response(401)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({"status": "error", "message": "Akses Ditolak! Anda bukan Admin."}).encode('utf-8'))
                return

            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            try:
                data = json.loads(post_data.decode('utf-8'))
                
                # 1. Check if Google Drive sync is configured
                if GOOGLE_SCRIPT_URL:
                    import urllib.request
                    
                    # Prepare monthly sheet name based on Date
                    date_str = data.get('tgl_terbit', '')
                    if not date_str or '-' not in date_str:
                        self.send_response(400)
                        self.send_header('Content-Type', 'application/json')
                        self.end_headers()
                        self.wfile.write(json.dumps({"status": "error", "message": "Format tanggal terbit tidak valid!"}).encode('utf-8'))
                        return
                    
                    pts = date_str.split('-')
                    year = pts[0]
                    month_idx = int(pts[1])
                    months_id = [
                        'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
                        'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember'
                    ]
                    sheet_name = f"{months_id[month_idx - 1]} {year}"
                    
                    payload = data.copy()
                    payload['action'] = 'add-ad'
                    payload['sheet_name'] = sheet_name
                    
                    # Call Google Apps Script Web App
                    req_data = json.dumps(payload).encode('utf-8')
                    req = urllib.request.Request(
                        GOOGLE_SCRIPT_URL,
                        data=req_data,
                        headers={'Content-Type': 'application/json', 'User-Agent': 'Mozilla/5.0'}
                    )
                    
                    with urllib.request.urlopen(req, timeout=10) as response:
                        res_json = json.loads(response.read().decode('utf-8'))
                        if res_json.get('status') == 'success':
                            # Trigger local cache update to keep in sync
                            try:
                                get_req = urllib.request.Request(GOOGLE_SCRIPT_URL, headers={'User-Agent': 'Mozilla/5.0'})
                                with urllib.request.urlopen(get_req, timeout=8) as get_res:
                                    live_json = get_res.read().decode('utf-8')
                                    parsed_data = json.loads(live_json)
                                    output_js = f"const INITIAL_DATA = {json.dumps(parsed_data, indent=2, ensure_ascii=False)};\n"
                                    output_js += "\n// Make available via window object just in case\nif (typeof window !== 'undefined') {\n  window.KOMPAS_DATA = INITIAL_DATA;\n}\n"
                                    with open('data.js', 'w', encoding='utf-8') as f:
                                        f.write(output_js)
                                    trigger_auto_git_push()
                            except Exception as cache_err:
                                print(f"Cache rebuild failed: {str(cache_err)}")
                            
                            self.send_response(200)
                            self.send_header('Content-Type', 'application/json')
                            self.end_headers()
                            self.wfile.write(json.dumps({"status": "success", "message": res_json.get('message')}).encode('utf-8'))
                            return
                        else:
                            raise Exception(res_json.get('message', 'Failed writing to Google Drive.'))
                
                # 2. Local Excel fallback
                success, msg = add_ad_to_excel(data)
                if success:
                    # Automatically trigger parsing compilation to rebuild data.js
                    generate_data.main()
                    trigger_auto_git_push()
                    self.send_response(200)
                    self.send_header('Content-Type', 'application/json')
                    self.end_headers()
                    self.wfile.write(json.dumps({"status": "success", "message": msg}).encode('utf-8'))
                else:
                    self.send_response(400)
                    self.send_header('Content-Type', 'application/json')
                    self.end_headers()
                    self.wfile.write(json.dumps({"status": "error", "message": msg}).encode('utf-8'))
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({"status": "error", "message": f"Server Error: {str(e)}"}).encode('utf-8'))
                
        # 3. API: Edit existing Ad negotiated Net Price
        elif path_clean == '/api/edit-price':
            # Check Admin Authentication header
            auth_header = self.headers.get('X-Admin-Password')
            if auth_header != ADMIN_PASSWORD:
                self.send_response(401)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({"status": "error", "message": "Akses Ditolak! Anda bukan Admin."}).encode('utf-8'))
                return

            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            try:
                data = json.loads(post_data.decode('utf-8'))
                so = data.get('so', '')
                price = data.get('price', '')
                sheet_name = data.get('sheet_name', '')
                row_idx = data.get('row_idx', None)
                
                if not sheet_name or (not so and not row_idx):
                    self.send_response(400)
                    self.send_header('Content-Type', 'application/json')
                    self.end_headers()
                    self.wfile.write(json.dumps({"status": "error", "message": "SO ID atau Baris tidak valid!"}).encode('utf-8'))
                    return
                    
                # 1. Check if Google Drive sync is configured
                if GOOGLE_SCRIPT_URL:
                    import urllib.request
                    
                    payload = {
                        'action': 'edit-price',
                        'sheet_name': sheet_name,
                        'so': so,
                        'price': price,
                        'row_idx': row_idx
                    }
                    
                    req_data = json.dumps(payload).encode('utf-8')
                    req = urllib.request.Request(
                        GOOGLE_SCRIPT_URL,
                        data=req_data,
                        headers={'Content-Type': 'application/json', 'User-Agent': 'Mozilla/5.0'}
                    )
                    
                    with urllib.request.urlopen(req, timeout=10) as response:
                        res_json = json.loads(response.read().decode('utf-8'))
                        if res_json.get('status') == 'success':
                            # Trigger local cache update to keep in sync
                            try:
                                get_req = urllib.request.Request(GOOGLE_SCRIPT_URL, headers={'User-Agent': 'Mozilla/5.0'})
                                with urllib.request.urlopen(get_req, timeout=8) as get_res:
                                    live_json = get_res.read().decode('utf-8')
                                    parsed_data = json.loads(live_json)
                                    output_js = f"const INITIAL_DATA = {json.dumps(parsed_data, indent=2, ensure_ascii=False)};\n"
                                    output_js += "\n// Make available via window object just in case\nif (typeof window !== 'undefined') {\n  window.KOMPAS_DATA = INITIAL_DATA;\n}\n"
                                    with open('data.js', 'w', encoding='utf-8') as f:
                                        f.write(output_js)
                                    trigger_auto_git_push()
                            except Exception as cache_err:
                                print(f"Cache rebuild failed: {str(cache_err)}")
                                
                            self.send_response(200)
                            self.send_header('Content-Type', 'application/json')
                            self.end_headers()
                            self.wfile.write(json.dumps({"status": "success", "message": res_json.get('message')}).encode('utf-8'))
                            return
                        else:
                            raise Exception(res_json.get('message', 'Failed writing price to Google Drive.'))
                            
                # 2. Local Excel fallback
                success, msg = edit_ad_price_in_excel(sheet_name, so, price, row_idx)
                
                if success:
                    generate_data.main()
                    trigger_auto_git_push()
                    self.send_response(200)
                    self.send_header('Content-Type', 'application/json')
                    self.end_headers()
                    self.wfile.write(json.dumps({"status": "success", "message": msg}).encode('utf-8'))
                else:
                    self.send_response(400)
                    self.send_header('Content-Type', 'application/json')
                    self.end_headers()
                    self.wfile.write(json.dumps({"status": "error", "message": msg}).encode('utf-8'))
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({"status": "error", "message": f"Server Error: {str(e)}"}).encode('utf-8'))
        
        # 3.5 API: Edit existing Ad Sales Order (SO) Number
        elif path_clean == '/api/edit-so':
            # Check Admin Authentication header
            auth_header = self.headers.get('X-Admin-Password')
            if auth_header != ADMIN_PASSWORD:
                self.send_response(401)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({"status": "error", "message": "Akses Ditolak! Anda bukan Admin."}).encode('utf-8'))
                return

            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            try:
                data = json.loads(post_data.decode('utf-8'))
                so = data.get('so', '')
                sheet_name = data.get('sheet_name', '')
                row_idx = data.get('row_idx', None)
                
                if not sheet_name or not row_idx:
                    self.send_response(400)
                    self.send_header('Content-Type', 'application/json')
                    self.end_headers()
                    self.wfile.write(json.dumps({"status": "error", "message": "Sheet atau Baris tidak valid!"}).encode('utf-8'))
                    return
                    
                # 1. Check if Google Drive sync is configured
                if GOOGLE_SCRIPT_URL:
                    import urllib.request
                    
                    payload = {
                        'action': 'edit-so',
                        'sheet_name': sheet_name,
                        'so': so,
                        'row_idx': row_idx
                    }
                    
                    req_data = json.dumps(payload).encode('utf-8')
                    req = urllib.request.Request(
                        GOOGLE_SCRIPT_URL,
                        data=req_data,
                        headers={'Content-Type': 'application/json', 'User-Agent': 'Mozilla/5.0'}
                    )
                    
                    with urllib.request.urlopen(req, timeout=10) as response:
                        res_json = json.loads(response.read().decode('utf-8'))
                        if res_json.get('status') == 'success':
                            # Trigger local cache update to keep in sync
                            try:
                                get_req = urllib.request.Request(GOOGLE_SCRIPT_URL, headers={'User-Agent': 'Mozilla/5.0'})
                                with urllib.request.urlopen(get_req, timeout=8) as get_res:
                                    live_json = get_res.read().decode('utf-8')
                                    parsed_data = json.loads(live_json)
                                    output_js = f"const INITIAL_DATA = {json.dumps(parsed_data, indent=2, ensure_ascii=False)};\n"
                                    output_js += "\n// Make available via window object just in case\nif (typeof window !== 'undefined') {\n  window.KOMPAS_DATA = INITIAL_DATA;\n}\n"
                                    with open('data.js', 'w', encoding='utf-8') as f:
                                        f.write(output_js)
                                    trigger_auto_git_push()
                            except Exception as cache_err:
                                print(f"Cache rebuild failed: {str(cache_err)}")
                                
                            self.send_response(200)
                            self.send_header('Content-Type', 'application/json')
                            self.end_headers()
                            self.wfile.write(json.dumps({"status": "success", "message": res_json.get('message')}).encode('utf-8'))
                            return
                        else:
                            raise Exception(res_json.get('message', 'Failed writing SO to Google Drive.'))
                            
                # 2. Local Excel fallback
                success, msg = edit_ad_so_in_excel(sheet_name, row_idx, so)
                
                if success:
                    generate_data.main()
                    trigger_auto_git_push()
                    self.send_response(200)
                    self.send_header('Content-Type', 'application/json')
                    self.end_headers()
                    self.wfile.write(json.dumps({"status": "success", "message": msg}).encode('utf-8'))
                else:
                    self.send_response(400)
                    self.send_header('Content-Type', 'application/json')
                    self.end_headers()
                    self.wfile.write(json.dumps({"status": "error", "message": msg}).encode('utf-8'))
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({"status": "error", "message": f"Server Error: {str(e)}"}).encode('utf-8'))
        
        # API: Edit existing Ad Notes (Keterangan/Link)
        elif path_clean == '/api/edit-notes':
            # Check Admin Authentication header
            auth_header = self.headers.get('X-Admin-Password')
            if auth_header != ADMIN_PASSWORD:
                self.send_response(401)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({"status": "error", "message": "Akses Ditolak! Anda bukan Admin."}).encode('utf-8'))
                return

            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            try:
                data = json.loads(post_data.decode('utf-8'))
                notes = data.get('notes', '')
                sheet_name = data.get('sheet_name', '')
                row_idx = data.get('row_idx', None)
                
                if not sheet_name or not row_idx:
                    self.send_response(400)
                    self.send_header('Content-Type', 'application/json')
                    self.end_headers()
                    self.wfile.write(json.dumps({"status": "error", "message": "Sheet atau Baris tidak valid!"}).encode('utf-8'))
                    return
                    
                # Local Excel fallback
                success, msg = edit_ad_notes_in_excel(sheet_name, row_idx, notes)
                
                if success:
                    generate_data.main()
                    trigger_auto_git_push()
                    self.send_response(200)
                    self.send_header('Content-Type', 'application/json')
                    self.end_headers()
                    self.wfile.write(json.dumps({"status": "success", "message": msg}).encode('utf-8'))
                else:
                    self.send_response(400)
                    self.send_header('Content-Type', 'application/json')
                    self.end_headers()
                    self.wfile.write(json.dumps({"status": "error", "message": msg}).encode('utf-8'))
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({"status": "error", "message": f"Server Error: {str(e)}"}).encode('utf-8'))
        else:
            self.send_response(404)
            self.end_headers()

    def do_GET(self):
        path_clean = self.path.split('?')[0]
        
        if path_clean == '/' or path_clean == '/index.html':
            self.log_visitor()
            
        # API: Get available memo prefixes
        if path_clean == '/api/available-memos':
            try:
                prefixes = []
                if os.path.exists(MEMO_DIR):
                    files = os.listdir(MEMO_DIR)
                    for f in files:
                        m = re.search(r'\b(\d+)\b', f)
                        if not m:
                            m = re.search(r'(\d+)', f)
                        if m:
                            prefixes.append(m.group(1))
                            prefixes.append(str(int(m.group(1))))
                
                # Deduplicate
                prefixes = list(set(prefixes))
                
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps(prefixes).encode('utf-8'))
            except Exception as e:
                self.send_response(500)
                self.end_headers()
                self.wfile.write(json.dumps([]).encode('utf-8'))
            return

        # API: Download memo file
        if path_clean == '/api/download-memo':
            from urllib.parse import urlparse, parse_qs
            query = parse_qs(urlparse(self.path).query)
            memo_val = query.get('memo', [''])[0]
            
            prefix = extract_memo_prefix(memo_val)
            if not prefix:
                self.send_response(400)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({"status": "error", "message": "Format nomor memo tidak dikenali!"}).encode('utf-8'))
                return
                
            file_path = match_memo_file(prefix)
            if not file_path or not os.path.exists(file_path):
                self.send_response(404)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({"status": "error", "message": f"Berkas memo untuk No. {prefix} tidak ditemukan di folder MEMO!"}).encode('utf-8'))
                return
                
            try:
                # Open and read file contents
                with open(file_path, 'rb') as f:
                    file_data = f.read()
                    
                # Determine Content-Type based on extension
                ext = os.path.splitext(file_path)[1].lower()
                content_type = 'application/octet-stream'
                if ext == '.pdf':
                    content_type = 'application/pdf'
                elif ext in ['.doc', '.docx']:
                    content_type = 'application/msword'
                elif ext in ['.xls', '.xlsx']:
                    content_type = 'application/vnd.ms-excel'
                elif ext in ['.jpg', '.jpeg']:
                    content_type = 'image/jpeg'
                elif ext == '.png':
                    content_type = 'image/png'
                    
                self.send_response(200)
                self.send_header('Content-Type', content_type)
                # Content-Disposition forces browser download
                filename = os.path.basename(file_path)
                self.send_header('Content-Disposition', f'attachment; filename="{filename}"')
                self.send_header('Content-Length', str(len(file_data)))
                self.end_headers()
                self.wfile.write(file_data)
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({"status": "error", "message": f"Gagal membaca file memo: {str(e)}"}).encode('utf-8'))
            return

        # API: Live Content - ambil data performa dari Google Sheets
        if path_clean == '/api/live-content':
            from urllib.parse import urlparse, parse_qs
            query = parse_qs(urlparse(self.path).query)
            sheet_name = query.get('sheet', [''])[0]
            if not sheet_name:
                # Coba auto-detect bulan berjalan
                import datetime
                months_id = ['Januari','Februari','Maret','April','Mei','Juni',
                             'Juli','Agustus','September','Oktober','November','Desember']
                now = datetime.datetime.now()
                sheet_name = f"{months_id[now.month - 1]} {now.year}"
            
            try:
                rows = fetch_live_content_csv(sheet_name)
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({
                    "sheet": sheet_name,
                    "rows": rows,
                    "count": len(rows)
                }, ensure_ascii=False).encode('utf-8'))
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({"error": str(e)}).encode('utf-8'))
            return

        # API: Auto-discover sheet GIDs dari Google Sheets
        if path_clean == '/api/live-sheet-gids':
            try:
                gids = fetch_all_live_sheet_gids()
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps(gids, ensure_ascii=False).encode('utf-8'))
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({"error": str(e)}).encode('utf-8'))
            return

        # API: Get visitor logs
        if path_clean == '/api/visitors':
            auth_header = self.headers.get('X-Admin-Password')
            if auth_header != ADMIN_PASSWORD:
                self.send_response(401)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({"status": "error", "message": "Akses Ditolak! Anda bukan Admin."}).encode('utf-8'))
                return
            
            try:
                if os.path.exists('visitors.json'):
                    with open('visitors.json', 'r', encoding='utf-8') as f:
                        data = f.read()
                else:
                    data = "[]"
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(data.encode('utf-8'))
            except Exception as e:
                self.send_response(500)
                self.end_headers()
                self.wfile.write(str(e).encode('utf-8'))
            return
            
        # API: Return the JSON dataset directly
        if path_clean == '/api/data':
            try:
                # 1. Check if Google Script URL is configured for live Drive syncing
                if GOOGLE_SCRIPT_URL:
                    import urllib.request
                    import urllib.parse
                    try:
                        req = urllib.request.Request(GOOGLE_SCRIPT_URL, headers={'User-Agent': 'Mozilla/5.0'})
                        with urllib.request.urlopen(req, timeout=8) as response:
                            if response.status == 200:
                                json_data = response.read().decode('utf-8')
                                # Save it locally to sync local cached file data.js
                                parsed_data = json.loads(json_data)
                                output_js = f"const INITIAL_DATA = {json.dumps(parsed_data, indent=2, ensure_ascii=False)};\n"
                                output_js += "\n// Make available via window object just in case\nif (typeof window !== 'undefined') {\n  window.KOMPAS_DATA = INITIAL_DATA;\n}\n"
                                with open('data.js', 'w', encoding='utf-8') as f:
                                    f.write(output_js)
                                
                                self.send_response(200)
                                self.send_header('Content-Type', 'application/json')
                                self.send_header('Access-Control-Allow-Origin', '*')
                                self.end_headers()
                                self.wfile.write(json_data.encode('utf-8'))
                                return
                    except Exception as drive_err:
                        print(f"Warning: Failed to fetch live Google Drive data ({str(drive_err)}). Falling back to local Excel cache.")

                # 2. Local Excel cache fallback
                # Load compiled data from data.js by reading the json block inside it
                # or simply run the parser script import
                if not os.path.exists('data.js'):
                    generate_data.main()
                
                with open('data.js', 'r', encoding='utf-8') as f:
                    content = f.read()
                
                # Parse JSON block inside data.js e.g., const/var INITIAL_DATA = {...};
                match = re.search(r'(?:const|var) INITIAL_DATA = ({.*?});', content, re.DOTALL)
                if match:
                    json_str = match.group(1)
                    self.send_response(200)
                    self.send_header('Content-Type', 'application/json')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.end_headers()
                    self.wfile.write(json_str.encode('utf-8'))
                else:
                    self.send_response(500)
                    self.end_headers()
            except Exception as e:
                self.send_response(500)
                self.end_headers()
                self.wfile.write(str(e).encode('utf-8'))
        else:
            # Standard static file serving
            super().do_GET()

def add_ad_to_excel(data):
    excel_path = 'Iklan Media Sosial Harian Kompas.xlsx'
    if not os.path.exists(excel_path):
        return False, "Berkas Excel master tidak ditemukan!"
    
    try:
        # Load in write mode (not read-only)
        wb = openpyxl.load_workbook(excel_path)
        
        # Determine the sheet name from raw Date (e.g. 2026-05-25 -> "Mei 2026")
        date_str = data.get('tgl_terbit', '')
        if not date_str or '-' not in date_str:
            return False, "Format tanggal terbit tidak valid!"
            
        pts = date_str.split('-')
        year = pts[0]
        month_idx = int(pts[1])
        
        months_id = [
            'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
            'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember'
        ]
        sheet_name = f"{months_id[month_idx - 1]} {year}"
        
        if sheet_name not in wb.sheetnames:
            # Create sheet if not exists, copying format from reference sheet
            ref_sheet = wb['Mei 2026'] if 'Mei 2026' in wb.sheetnames else wb.active
            sheet = wb.create_sheet(title=sheet_name)
            
            # Copy headers row 1 (MEDIA SOSIAL) and row 2 (headers)
            for r in range(1, 3):
                for c in range(1, len(ref_sheet[r]) + 1):
                    sheet.cell(row=r, column=c, value=ref_sheet.cell(row=r, column=c).value)
        else:
            sheet = wb[sheet_name]
            
        # Find next available row by scanning
        next_row = 3
        while sheet.cell(row=next_row, column=2).value is not None or sheet.cell(row=next_row, column=3).value is not None or sheet.cell(row=next_row, column=9).value is not None:
            next_row += 1
            
        # Determine No.
        no_val = 1
        for r in range(3, next_row):
            val = sheet.cell(row=r, column=1).value
            if val is not None:
                try:
                    no_val = max(no_val, int(float(val)) + 1)
                except:
                    pass
                    
        # Set values
        # Column mapping: NO, JUDUL IKLAN, POSISI, TOTAL AD, TGL. TERBIT, AE/CP, KETERANGAN ORDER, Keterangan, Sales Order (SO)
        day_num = int(pts[2])
        month_short = months_id[month_idx - 1][:3] # e.g., "Mei", "Apr", "May"
        
        # Language adjustments for month name to match snapshot format
        month_mapping = {
            'Januari': 'Jan', 'Februari': 'Feb', 'Maret': 'Mar', 'April': 'Apr',
            'Mei': 'Mei', 'Juni': 'Juni', 'Juli': 'Juli', 'Agustus': 'Agustus',
            'September': 'Sept', 'Oktober': 'Okt', 'November': 'Nov', 'Desember': 'Des'
        }
        month_display = month_mapping.get(months_id[month_idx - 1], month_short)
        
        tgl_terbit_str = f"{str(day_num).zfill(2)} {month_display}"
        
        sheet.cell(row=next_row, column=1, value=no_val)
        sheet.cell(row=next_row, column=2, value=data.get('judul', ''))
        sheet.cell(row=next_row, column=3, value=data.get('posisi', ''))
        sheet.cell(row=next_row, column=4, value=int(data.get('total_ad', 1)))
        sheet.cell(row=next_row, column=5, value=tgl_terbit_str)
        sheet.cell(row=next_row, column=6, value=data.get('ae', ''))
        sheet.cell(row=next_row, column=7, value=data.get('keterangan_order', ''))
        sheet.cell(row=next_row, column=8, value=data.get('keterangan', ''))
        sheet.cell(row=next_row, column=9, value=data.get('so', ''))
        
        # Column 10 (J) is PENDAPATAN
        custom_price = data.get('pendapatan', '')
        if custom_price:
            try:
                sheet.cell(row=next_row, column=10, value=int(float(custom_price)))
            except:
                sheet.cell(row=next_row, column=10, value=custom_price)
        
        wb.save(excel_path)
        return True, f"Sukses menambahkan baris iklan #{no_val} di sheet '{sheet_name}'"
    except PermissionError:
        return False, "⚠️ File Excel sedang terbuka! Tutup 'Iklan Media Sosial Harian Kompas.xlsx' di Microsoft Excel, lalu coba lagi."
    except Exception as e:
        return False, f"Gagal memproses penulisan Excel: {str(e)}"

def edit_ad_price_in_excel(sheet_name, so, price, row_idx=None):
    excel_path = 'Iklan Media Sosial Harian Kompas.xlsx'
    if not os.path.exists(excel_path):
        return False, "Berkas Excel master tidak ditemukan!"
        
    try:
        wb = openpyxl.load_workbook(excel_path)
        if sheet_name not in wb.sheetnames:
            return False, f"Sheet '{sheet_name}' tidak ditemukan di berkas Excel!"
            
        sheet = wb[sheet_name]
        found_row = None
        
        # 1. Try to identify by row_idx if provided
        if row_idx is not None:
            try:
                r = int(row_idx)
                if 3 <= r <= sheet.max_row:
                    found_row = r
            except Exception as ex:
                print("Error parsing row_idx:", ex)
                
        # 2. Fallback to scanning Column I (Sales Order - SO) if row not found
        if not found_row and so:
            target_so = str(so).strip()
            for r in range(3, sheet.max_row + 1):
                cell_val = sheet.cell(row=r, column=9).value
                if cell_val is not None:
                    cell_so = str(cell_val).strip()
                    if cell_so == target_so:
                        found_row = r
                        break
                        
        if not found_row:
            if so:
                return False, f"Sales Order (SO) ID '{so}' tidak ditemukan di sheet '{sheet_name}'!"
            else:
                return False, f"Baris indeks '{row_idx}' tidak valid di sheet '{sheet_name}'!"
            
        # Ensure Column J has header "PENDAPATAN" on row 2 if it's empty
        if sheet.cell(row=2, column=10).value is None:
            sheet.cell(row=2, column=10, value="PENDAPATAN")
            
        if price == "" or price is None:
            sheet.cell(row=found_row, column=10, value=None)
            msg = f"Sukses mereset harga baris {found_row} ke tarif resmi Rate Card."
        else:
            try:
                # Clean formatting like dots, commas, spaces, and "Rp" if price is string
                if isinstance(price, str) and price.strip() != "":
                    cleaned = re.sub(r'[^\d]', '', price)
                    price_val = int(cleaned)
                else:
                    price_val = int(float(price))
                sheet.cell(row=found_row, column=10, value=price_val)
                msg = f"Sukses mengubah harga baris {found_row} menjadi Rp {price_val:,}."
            except:
                sheet.cell(row=found_row, column=10, value=price)
                msg = f"Sukses mengubah harga baris {found_row} menjadi {price}."
                
        wb.save(excel_path)
        return True, msg
    except PermissionError:
        return False, "⚠️ File Excel sedang terbuka! Tutup 'Iklan Media Sosial Harian Kompas.xlsx' di Microsoft Excel, lalu coba lagi."
    except Exception as e:
        return False, f"Gagal mengubah harga di Excel: {str(e)}"

def edit_ad_so_in_excel(sheet_name, row_idx, new_so):
    excel_path = 'Iklan Media Sosial Harian Kompas.xlsx'
    if not os.path.exists(excel_path):
        return False, "Berkas Excel master tidak ditemukan!"
    try:
        wb = openpyxl.load_workbook(excel_path)
        if sheet_name not in wb.sheetnames:
            return False, f"Sheet '{sheet_name}' tidak ditemukan di berkas Excel!"
        sheet = wb[sheet_name]
        r = int(row_idx)
        if r < 3 or r > sheet.max_row:
            return False, f"Baris indeks '{row_idx}' tidak valid!"
        
        # Column 9 (I) is the Sales Order (SO) column
        sheet.cell(row=r, column=9, value=str(new_so).strip())
        wb.save(excel_path)
        return True, f"Sukses mengubah Nomor SO baris {row_idx} menjadi '{new_so}'."
    except PermissionError:
        return False, "⚠️ File Excel sedang terbuka! Tutup 'Iklan Media Sosial Harian Kompas.xlsx' di Microsoft Excel, lalu coba lagi."
    except Exception as e:
        return False, f"Gagal mengubah Nomor SO di Excel: {str(e)}"

def edit_ad_notes_in_excel(sheet_name, row_idx, new_notes):
    excel_path = 'Iklan Media Sosial Harian Kompas.xlsx'
    if not os.path.exists(excel_path):
        return False, "Berkas Excel master tidak ditemukan!"
    try:
        wb = openpyxl.load_workbook(excel_path)
        if sheet_name not in wb.sheetnames:
            return False, f"Sheet '{sheet_name}' tidak ditemukan di berkas Excel!"
        sheet = wb[sheet_name]
        r = int(row_idx)
        if r < 3 or r > sheet.max_row:
            return False, f"Baris indeks '{row_idx}' tidak valid!"
        
        # Column 8 (H) is the Keterangan column
        sheet.cell(row=r, column=8, value=str(new_notes).strip())
        wb.save(excel_path)
        return True, f"Sukses mengubah Keterangan baris {row_idx}."
    except PermissionError:
        return False, (
            "⚠️ File Excel sedang terbuka di Microsoft Excel!\n"
            "Tutup terlebih dahulu file 'Iklan Media Sosial Harian Kompas.xlsx' "
            "di Excel, lalu coba simpan kembali."
        )
    except Exception as e:
        return False, f"Gagal mengubah Keterangan di Excel: {str(e)}"

if __name__ == "__main__":
    # Ensure dataset is generated and synced with Google Drive on startup ONLY if local file is missing
    excel_path = 'Iklan Media Sosial Harian Kompas.xlsx'
    sync_on_start = not os.path.exists(excel_path)
    if sync_on_start:
        print("Berkas Excel lokal tidak ditemukan. Mempersiapkan server dan mendownload data awal...")
    else:
        print("Mempersiapkan server menggunakan berkas Excel lokal yang sudah ada...")
    generate_data.main(sync=sync_on_start)
        
    handler = CustomHandler
    # Enable socket reuse
    socketserver.TCPServer.allow_reuse_address = True

    # Subclass TCPServer to suppress benign connection errors from browser
    class QuietTCPServer(socketserver.TCPServer):
        def handle_error(self, request, client_address):
            import traceback, sys
            exc = sys.exc_info()[1]
            # Suppress WinError 10053/10054 (browser closed connection) — these are harmless
            if isinstance(exc, (ConnectionAbortedError, ConnectionResetError, BrokenPipeError)):
                return
            if hasattr(exc, 'winerror') and exc.winerror in (10053, 10054):
                return
            # Log other real errors normally
            print(f"[server-error] Exception dari {client_address}:")
            traceback.print_exc()
    
    with QuietTCPServer(("", PORT), handler) as httpd:
        print(f"Backend Server master berjalan di http://localhost:{PORT}")
        print("Buka browser dan buka tautan di atas untuk menggunakan dashboard.")
        try:
            httpd.serve_forever()
        except KeyboardInterrupt:
            print("\nServer dihentikan.")

