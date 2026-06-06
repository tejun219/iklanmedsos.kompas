import openpyxl
import json
import datetime
import re
import os

def parse_date(date_val, sheet_name):
    if isinstance(date_val, datetime.datetime):
        return date_val.strftime('%Y-%m-%d')
    if isinstance(date_val, (int, float)):
        # If it's a number, it could be a day number or serial date
        # Let's see if we can convert it to day number
        try:
            day = int(date_val)
            # parse sheet_name like "Mei 2026"
            sheet_match = re.match(r'([a-zA-Z]+)\s+(\d{4})', sheet_name)
            if sheet_match:
                year = int(sheet_match.group(2))
                month_str = sheet_match.group(1).lower()
                
                months_id = {
                    'januari': 1, 'februari': 2, 'maret': 3, 'april': 4, 'mei': 5, 'juni': 6,
                    'juli': 7, 'agustus': 8, 'september': 9, 'oktober': 10, 'november': 11, 'desember': 12,
                    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6, 'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
                }
                month = months_id.get(month_str, 5)
                dt = datetime.datetime(year, month, day)
                return dt.strftime('%Y-%m-%d')
        except:
            pass
        return str(int(date_val)) if isinstance(date_val, float) else str(date_val)
    if not date_val:
        return ""
    
    # Try parsing string like "08 Mei", "02 May", "2 May", "25 Mei"
    date_str = str(date_val).strip()
    match = re.match(r'(\d+)\s+([a-zA-Z]+)', date_str)
    if match:
        day = int(match.group(1))
        month_str = match.group(2).lower()
        
        # Determine year and month from sheet name e.g. "Mei 2026"
        sheet_match = re.match(r'([a-zA-Z]+)\s+(\d{4})', sheet_name)
        if sheet_match:
            year = int(sheet_match.group(2))
        else:
            year = 2026 # fallback
            
        months_id = {
            'januari': 1, 'februari': 2, 'maret': 3, 'april': 4, 'mei': 5, 'juni': 6,
            'juli': 7, 'agustus': 8, 'september': 9, 'oktober': 10, 'november': 11, 'desember': 12,
            'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6, 'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
        }
        
        month = months_id.get(month_str, 5) # default to May/Mei
        if month_str == 'may':
            month = 5
            
        try:
            dt = datetime.datetime(year, month, day)
            return dt.strftime('%Y-%m-%d')
        except ValueError:
            pass
            
    return date_str

def sync_google_drive(excel_path):
    import urllib.request
    url = 'https://docs.google.com/spreadsheets/d/1HLsXPCyyV4_evycJdmw4U9JrUHSwcyva/export?format=xlsx'
    try:
        print("Mendownload data terbaru dari Google Drive...")
        urllib.request.urlretrieve(url, excel_path)
        print("Berhasil mengunduh data dari Google Drive.")
        return True
    except Exception as e:
        print(f"Gagal mengunduh dari Google Drive: {e}")
        return False

def main(sync=False):
    excel_path = 'Iklan Media Sosial Harian Kompas.xlsx'
    
    if sync:
        sync_google_drive(excel_path)
        
    if not os.path.exists(excel_path):
        print(f"Error: {excel_path} not found.")
        return

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    all_data = {}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        
        # Determine headers row
        # In typical sheets, Row 1 has "MEDIA SOSIAL", Row 2 has the headers
        raw_headers = []
        data_start_row = 2
        
        # Let's inspect first few rows to find headers
        found_headers = False
        for idx in range(min(5, len(rows))):
            row_vals = [str(x).strip().upper() if x is not None else "" for x in rows[idx]]
            if any("JUDUL" in x for x in row_vals):
                raw_headers = [str(x).strip() if x is not None else "" for x in rows[idx]]
                data_start_row = idx + 1
                found_headers = True
                break
                
        if not found_headers:
            print(f"Skipping sheet {sheet_name}: Headers not found.")
            continue
            
        # Clean headers
        header_map = []
        for h in raw_headers:
            hu = h.upper()
            if "NO" == hu:
                header_map.append("no")
            elif "JUDUL" in hu:
                header_map.append("judul")
            elif "POSISI" in hu:
                header_map.append("posisi")
            elif "TOTAL" in hu:
                header_map.append("total_ad")
            elif "TGL" in hu or "TANGGAL" in hu:
                header_map.append("tgl_terbit")
            elif "AE" in hu or "CP" in hu:
                header_map.append("ae")
            elif "ORDER" in hu and "SO" not in hu:
                header_map.append("keterangan_order")
            elif "KETERANGAN" == hu or "KETERANGAN" in hu:
                header_map.append("keterangan")
            elif "SO" in hu or "SALES" in hu:
                header_map.append("so")
            elif "PENDAPATAN" in hu or "HARGA" in hu:
                header_map.append("pendapatan")
            else:
                header_map.append(h.lower().replace(" ", "_") if h else "unknown")
                
        sheet_items = []
        for r_idx in range(data_start_row, len(rows)):
            row = rows[r_idx]
            if all(cell is None for cell in row):
                continue
            
            item = {}
            for h_idx, key in enumerate(header_map):
                if key == "unknown" or h_idx >= len(row):
                    continue
                val = row[h_idx]
                if key == "tgl_terbit":
                    val = parse_date(val, sheet_name)
                elif key == "no" and val is not None:
                    try:
                        val = int(float(val))
                    except:
                        pass
                elif key == "total_ad" and val is not None:
                    try:
                        val = int(float(val))
                    except:
                        val = 1
                elif key == "pendapatan" and val is not None:
                    try:
                        val = int(float(val))
                    except:
                        val = ""
                
                # Format string cleanups
                if isinstance(val, str):
                    val = val.strip()
                item[key] = val if val is not None else ""
                
            # Add the raw Excel row index (1-indexed)
            item['row_idx'] = r_idx + 1

            # Filter out empty records
            if not item.get("judul") and not item.get("posisi") and not item.get("so"):
                continue
                
            sheet_items.append(item)
            
        all_data[sheet_name] = sheet_items
        print(f"Parsed sheet {sheet_name}: {len(sheet_items)} rows")

    # Generate MEMO_MAP
    memo_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "memo")
    memo_map = {}
    if os.path.exists(memo_dir):
        for f in os.listdir(memo_dir):
            m = re.match(r'^(\d+)', f)
            if m:
                prefix = m.group(1)
                memo_map[prefix] = f
                memo_map[str(int(prefix))] = f

    # Generate data.js containing the data
    output_js = f"var INITIAL_DATA = {json.dumps(all_data, indent=2, ensure_ascii=False)};\n"
    output_js += f"var MEMO_MAP = {json.dumps(memo_map, indent=2, ensure_ascii=False)};\n"
    output_js += "\n// Make available via window object for dynamic script loading compatibility\nif (typeof window !== 'undefined') {\n  window.INITIAL_DATA = INITIAL_DATA;\n  window.KOMPAS_DATA = INITIAL_DATA;\n  window.MEMO_MAP = MEMO_MAP;\n}\n"
    
    with open('data.js', 'w', encoding='utf-8') as f:
        f.write(output_js)
    print("Done parsing Excel! Written to data.js")

if __name__ == "__main__":
    main()
